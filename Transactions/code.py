import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('transactions.xlsx')

#specify sheet
sheet = wb['Sheet1']

#access cell using coordinates(row,column)
cell = sheet.cell(1, 1)

#get # of rows
print(sheet.max_row)

print(cell.value)

# +1 to include last one
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)

    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet, 
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

BarChart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')


#saves new file 
wb.save('transactions2.xlsx')